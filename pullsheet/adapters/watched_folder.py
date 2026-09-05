"""The primary ingestion path and the demo centrepiece.

A district exports inventory to a network folder on a schedule. Nobody logs into
anything; the file simply appears. This adapter polls that folder, reads CSV and
XLSX through ``column_map``, and moves each processed file to ``data/archive/``.

**Archive on success only.** A rejected file stays exactly where it landed, so
the person who dropped it can see that it is still there. A rejection that
quietly tidies the evidence away is worse than no rejection at all.
"""

from __future__ import annotations

import csv
import shutil
from pathlib import Path
from typing import Iterator

from pullsheet.adapters.base import AdapterRejection, InventoryAdapter, NormalizedRecord
from pullsheet.adapters.column_map import apply, detect, required_missing

ROOT = Path(__file__).resolve().parent.parent.parent
WATCHED = ROOT / "data" / "watched"
ARCHIVE = ROOT / "data" / "archive"

READABLE_SUFFIXES = {".csv", ".xlsx", ".xlsm"}


def _digits(value: str | None) -> str | None:
    if not value:
        return None
    kept = "".join(c for c in str(value) if c.isdigit())
    return kept or None


def _number(value: str | None) -> float | None:
    """Parse a number, or return None. Never guesses.

    A blank quantity stays blank. Defaulting it to 1 would invent a case of food
    that may not exist, and the operator would have no way to tell.
    """
    if value is None:
        return None
    text = str(value).strip().replace(",", "").replace("$", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


class WatchedFolderAdapter(InventoryAdapter):
    name = "watched_folder"
    provenance = "live"

    def declares(self) -> frozenset[str]:
        """Everything a district export normally carries. Honest by construction:
        anything a given file omits comes back in ``unpopulated`` per row."""
        return frozenset({
            "site", "storage_location", "raw_description", "quantity", "unit",
            "pack_size", "gtin", "upc", "lot_code", "unit_cost", "received_date",
        })

    # -- reading -----------------------------------------------------------

    def read(self, source, column_map: dict[str, str] | None = None
             ) -> Iterator[NormalizedRecord]:
        path = Path(source)
        rows, headers = self._rows(path)

        if not headers:
            raise AdapterRejection(path.name, None, "no header row; the file is empty")

        mapping = column_map or detect(headers)[0]
        missing = required_missing(mapping)
        if missing:
            raise AdapterRejection(
                path.name, ", ".join(sorted(missing)),
                f"no column matched the required field(s): {', '.join(sorted(missing))}. "
                f"Headers seen: {', '.join(headers)}")

        for source_row, row in enumerate(rows, start=1):
            self._guard_against_broken_quoting(path, source_row, row)
            yield self._record(mapping, row, source_row)

    @staticmethod
    def _guard_against_broken_quoting(path: Path, source_row: int, row: dict) -> None:
        """Unterminated quotes make csv swallow later lines into one field.

        The signature is a newline inside a value: no real inventory description
        contains one. Rejecting the whole source is right here -- half a file
        looks exactly like a district with fewer items in it.
        """
        for header, value in row.items():
            if isinstance(value, str) and "\n" in value:
                raise AdapterRejection(
                    path.name, f"row {source_row}, column {header!r}",
                    "unterminated quote: the value runs across a line break, so "
                    "the rest of the file was absorbed into one cell")

    def _rows(self, path: Path) -> tuple[list[dict], list[str] | None]:
        if path.suffix.lower() == ".csv":
            with path.open(newline="") as f:
                reader = csv.DictReader(f)
                return list(reader), reader.fieldnames
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            wb.close()
            if not rows:
                return [], None
            headers = [str(h) if h is not None else "" for h in rows[0]]
            body = [
                {headers[i]: ("" if v is None else str(v)) for i, v in enumerate(r[:len(headers)])}
                for r in rows[1:]
                if any(v is not None and str(v).strip() for v in r)
            ]
            return body, headers
        raise AdapterRejection(path.name, None,
                               f"unsupported file type {path.suffix!r}; expected .csv or .xlsx")

    @staticmethod
    def _record(mapping: dict[str, str], row: dict, source_row: int) -> NormalizedRecord:
        f = apply(mapping, row)
        unpopulated: set[str] = set()

        def keep(field: str, value):
            if value in (None, ""):
                unpopulated.add(field)
                return None
            return value

        site = (f.get("site") or "").strip()
        if not site:
            unpopulated.add("site")
        # raw_description is stored even when blank: FR-007 forbids dropping a
        # row we could not read, and an empty description with the field flagged
        # is a row an operator can still go and look at.
        description = (f.get("raw_description") or "").strip()
        if not description:
            unpopulated.add("raw_description")

        quantity = _number(f.get("quantity"))
        if quantity is None:
            unpopulated.add("quantity")

        gtin = _digits(f.get("gtin"))
        if not gtin:
            unpopulated.add("gtin")

        unit_cost = _number(f.get("unit_cost"))
        if unit_cost is None:
            unpopulated.add("unit_cost")

        return NormalizedRecord(
            site=site,
            storage_location=keep("storage_location", (f.get("storage_location") or "").strip() or None),
            raw_description=description,
            quantity=quantity,
            unit=keep("unit", (f.get("unit") or "").strip() or None),
            pack_size=keep("pack_size", (f.get("pack_size") or "").strip() or None),
            gtin=gtin,
            upc=gtin,
            # VERBATIM. Case, punctuation and whitespace exactly as written (R3).
            lot_code=keep("lot_code", f.get("lot_code") or None),
            unit_cost=unit_cost,
            received_date=keep("received_date", (f.get("received_date") or "").strip() or None),
            source_row=source_row,
            unpopulated=frozenset(unpopulated),
        )

    # -- polling -----------------------------------------------------------

    @staticmethod
    def pending(folder: Path = WATCHED) -> list[Path]:
        """Files waiting to be ingested, oldest first."""
        if not folder.exists():
            return []
        return sorted(
            (p for p in folder.iterdir()
             if p.is_file() and not p.name.startswith(".")),
            key=lambda p: p.stat().st_mtime,
        )

    @staticmethod
    def archive(path: Path, archive_dir: Path = ARCHIVE) -> Path:
        """Move a successfully ingested file out of the watched folder."""
        archive_dir.mkdir(parents=True, exist_ok=True)
        target = archive_dir / path.name
        if target.exists():
            target = archive_dir / f"{path.stem}-{int(path.stat().st_mtime)}{path.suffix}"
        shutil.move(str(path), str(target))
        return target


# ---------------------------------------------------------------------------
# The poll loop (T035)
# ---------------------------------------------------------------------------

def poll_once(db_path=None, folder: Path = WATCHED, archive_dir: Path = ARCHIVE) -> list[dict]:
    """Ingest every file waiting in the watched folder, then re-run the matcher.

    Archive on SUCCESS only. A rejected file stays where it landed so the person
    who dropped it can see that it is still there -- a rejection that tidies away
    its own evidence is worse than no rejection at all.
    """
    from pullsheet import db as db_module
    from pullsheet.matching.run import run_matcher

    results: list[dict] = []
    pending = WatchedFolderAdapter.pending(folder)
    if not pending:
        return results

    conn = db_module.connect(db_path or db_module.DB_PATH)
    adapter = WatchedFolderAdapter()
    try:
        for path in pending:
            if path.suffix.lower() not in READABLE_SUFFIXES:
                results.append({"status": "rejected", "filename": path.name,
                                "reason": f"unsupported file type {path.suffix!r}"})
                continue
            outcome = db_module.ingest_file(conn, path, adapter,
                                            f"{path.parent.name} watched folder")
            if outcome["status"] == "ok":
                WatchedFolderAdapter.archive(path, archive_dir)
                outcome["archived"] = True
            results.append(outcome)

        if any(r["status"] == "ok" for r in results):
            run_matcher(conn)
    finally:
        conn.close()
    return results


def watch(interval_seconds: float = 2.0, stop=None, db_path=None) -> None:
    """Poll forever. Intended to run on a daemon thread beside uvicorn.

    Exceptions are swallowed deliberately: a poller that dies stops watching the
    folder, and a folder nobody is watching looks exactly like a district with
    nothing recalled.
    """
    import time
    import traceback

    while stop is None or not stop.is_set():
        try:
            for result in poll_once(db_path=db_path):
                print(f"[watched-folder] {result['status']}: {result['filename']}"
                      + (f" -- {result.get('reason', '')}" if result["status"] == "rejected" else ""))
        except Exception:                      # noqa: BLE001
            traceback.print_exc()
        time.sleep(interval_seconds)
