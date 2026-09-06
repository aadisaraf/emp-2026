"""The primary ingestion path and the demo centrepiece.

The location's inventory software writes an export to an SFTP drop on a
schedule, once a day. Nobody logs into anything; the file simply appears. This
adapter watches the drop directory, reads CSV and XLSX through ``column_map``,
and moves each processed file to ``data/archive/``.

It is a directory watcher, not an SFTP client: the SFTP server writes into the
directory and this reads it. That is the same shape a real deployment has, and
it keeps the credential handling outside the application entirely.

**Archive on success only.** A rejected file stays exactly where it landed, so
the person who dropped it can see that it is still there. A rejection that
quietly tidies the evidence away is worse than no rejection at all.

**Only settled files are read.** An SFTP write is not atomic, so a file still
being uploaded can be read as a complete short export -- indistinguishable from
a kitchen that genuinely holds less food. ``pending()`` skips anything modified
in the last few seconds.
"""

from __future__ import annotations

import csv
import shutil
from pathlib import Path
from typing import Iterator

from pullsheet.adapters.base import AdapterRejection, InventoryAdapter, NormalizedRecord
from pullsheet.adapters.column_map import apply, detect, required_missing

ROOT = Path(__file__).resolve().parent.parent.parent
WATCHED = ROOT / "data" / "watched"
ARCHIVE = ROOT / "data" / "archive"

READABLE_SUFFIXES = {".csv", ".xlsx", ".xlsm"}


def _digits(value: str | None) -> str | None:
    if not value:
        return None
    kept = "".join(c for c in str(value) if c.isdigit())
    return kept or None


def _number(value: str | None) -> float | None:
    """Parse a number, or return None. Never guesses.

    A blank quantity stays blank. Defaulting it to 1 would invent a case of food
    that may not exist, and the operator would have no way to tell.
    """
    if value is None:
        return None
    text = str(value).strip().replace(",", "").replace("$", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


class SftpDropAdapter(InventoryAdapter):
    name = "sftp_drop"
    provenance = "live"
    channel = "sftp_drop"

    def declares(self) -> frozenset[str]:
        """Everything an inventory export normally carries. Honest by construction:
        anything a given file omits comes back in ``unpopulated`` per row."""
        return frozenset({
            "storage_location", "raw_description", "quantity", "unit",
            "pack_size", "gtin", "lot_code", "brand", "manufacturer",
            "manufacturer_item_code", "vendor_name", "vendor_item_code",
            "unit_cost", "received_date",
        })

    # -- reading -----------------------------------------------------------

    def read(self, source, column_map: dict[str, str] | None = None
             ) -> Iterator[NormalizedRecord]:
        path = Path(source)
        rows, headers = self._rows(path)

        if not headers:
            raise AdapterRejection(path.name, None, "no header row; the file is empty")

        mapping = column_map or detect(headers)[0]
        missing = required_missing(mapping)
        if missing:
            raise AdapterRejection(
                path.name, ", ".join(sorted(missing)),
                f"no column matched the required field(s): {', '.join(sorted(missing))}. "
                f"Headers seen: {', '.join(headers)}")

        for source_row, row in enumerate(rows, start=1):
            self._guard_against_broken_quoting(path, source_row, row)
            yield self._record(mapping, row, source_row)

    @staticmethod
    def _guard_against_broken_quoting(path: Path, source_row: int, row: dict) -> None:
        """Unterminated quotes make csv swallow later lines into one field.

        The signature is a newline inside a value: no real inventory description
        contains one. Rejecting the whole source is right here -- half a file
        looks exactly like a district with fewer items in it.
        """
        for header, value in row.items():
            if isinstance(value, str) and "\n" in value:
                raise AdapterRejection(
                    path.name, f"row {source_row}, column {header!r}",
                    "unterminated quote: the value runs across a line break, so "
                    "the rest of the file was absorbed into one cell")

    def _rows(self, path: Path) -> tuple[list[dict], list[str] | None]:
        if path.suffix.lower() == ".csv":
            with path.open(newline="") as f:
                reader = csv.DictReader(f)
                return list(reader), reader.fieldnames
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            wb.close()
            if not rows:
                return [], None
            headers = [str(h) if h is not None else "" for h in rows[0]]
            body = [
                {headers[i]: ("" if v is None else str(v)) for i, v in enumerate(r[:len(headers)])}
                for r in rows[1:]
                if any(v is not None and str(v).strip() for v in r)
            ]
            return body, headers
        raise AdapterRejection(path.name, None,
                               f"unsupported file type {path.suffix!r}; expected .csv or .xlsx")

    @staticmethod
    def _record(mapping: dict[str, str], row: dict, source_row: int) -> NormalizedRecord:
        f = apply(mapping, row)
        unpopulated: set[str] = set()

        def keep(field: str, value):
            if value in (None, ""):
                unpopulated.add(field)
                return None
            return value

        # raw_description is stored even when blank: FR-007 forbids dropping a
        # row we could not read, and an empty description with the field flagged
        # is a row an operator can still go and look at.
        description = (f.get("raw_description") or "").strip()
        if not description:
            unpopulated.add("raw_description")

        quantity = _number(f.get("quantity"))
        if quantity is None:
            unpopulated.add("quantity")

        gtin = _digits(f.get("gtin"))
        if not gtin:
            unpopulated.add("gtin")

        unit_cost = _number(f.get("unit_cost"))
        if unit_cost is None:
            unpopulated.add("unit_cost")

        # Supplier identity (FR-069). Passed through as the source wrote it:
        # firm.agrees() does the normalizing, in one place, so an adapter cannot
        # change what matches by tidying a company name differently.
        supplier = {name: keep(name, (f.get(name) or "").strip() or None)
                    for name in ("brand", "manufacturer", "manufacturer_item_code",
                                 "vendor_name", "vendor_item_code")}

        return NormalizedRecord(
            storage_location=keep("storage_location", (f.get("storage_location") or "").strip() or None),
            raw_description=description,
            quantity=quantity,
            unit=keep("unit", (f.get("unit") or "").strip() or None),
            pack_size=keep("pack_size", (f.get("pack_size") or "").strip() or None),
            gtin=gtin,
            # VERBATIM. Case, punctuation and whitespace exactly as written (R3).
            lot_code=keep("lot_code", f.get("lot_code") or None),
            **supplier,
            unit_cost=unit_cost,
            received_date=keep("received_date", (f.get("received_date") or "").strip() or None),
            source_row=source_row,
            unpopulated=frozenset(unpopulated),
        )

    # -- polling -----------------------------------------------------------

    #: Seconds a file must sit unchanged before it is considered fully written.
    #: An SFTP upload lands byte by byte, and a CSV cut off on a row boundary
    #: parses perfectly as a shorter inventory -- a partial read is the one
    #: failure here that produces a plausible wrong answer instead of an error.
    SETTLE_SECONDS = 2.0

    @classmethod
    def pending(cls, folder: Path = WATCHED, now: float | None = None) -> list[Path]:
        """Files that have finished arriving, oldest first."""
        if not folder.exists():
            return []
        import time

        now = time.time() if now is None else now
        return sorted(
            (p for p in folder.iterdir()
             if p.is_file() and not p.name.startswith(".")
             and now - p.stat().st_mtime >= cls.SETTLE_SECONDS),
            key=lambda p: p.stat().st_mtime,
        )

    @staticmethod
    def archive(path: Path, archive_dir: Path = ARCHIVE) -> Path:
        """Move a successfully ingested file out of the watched folder."""
        archive_dir.mkdir(parents=True, exist_ok=True)
        target = archive_dir / path.name
        if target.exists():
            target = archive_dir / f"{path.stem}-{int(path.stat().st_mtime)}{path.suffix}"
        shutil.move(str(path), str(target))
        return target


# ---------------------------------------------------------------------------
# The poll loop (T035)
# ---------------------------------------------------------------------------

def poll_once(db_path=None, folder: Path = WATCHED, archive_dir: Path = ARCHIVE) -> list[dict]:
    """Ingest every settled file in the drop directory as its own run.

    Each file is one run, carried all the way through matching and finalize by
    ``db.ingest_file`` -- so a crash cannot leave rows committed with no matches
    and a sheet that reads as good news.

    Archive on SUCCESS only. A rejected file stays where it landed so the person
    who dropped it can see that it is still there -- a rejection that tidies away
    its own evidence is worse than no rejection at all.
    """
    from pullsheet import db as db_module

    results: list[dict] = []
    pending = SftpDropAdapter.pending(folder)
    if not pending:
        return results

    conn = db_module.connect(db_path or db_module.DB_PATH)
    adapter = SftpDropAdapter()
    try:
        for path in pending:
            if path.suffix.lower() not in READABLE_SUFFIXES:
                results.append({"status": "rejected", "filename": path.name,
                                "reason": f"unsupported file type {path.suffix!r}"})
                continue
            outcome = db_module.ingest_file(conn, path, adapter)
            if outcome["status"] in ("ok", "duplicate"):
                SftpDropAdapter.archive(path, archive_dir)
                outcome["archived"] = True
            results.append(outcome)
    finally:
        conn.close()
    return results


def watch(interval_seconds: float = 2.0, stop=None, db_path=None) -> None:
    """Poll forever. Intended to run on a daemon thread beside uvicorn.

    Exceptions are swallowed deliberately: a poller that dies stops watching the
    drop, and a drop nobody is watching looks exactly like a kitchen with
    nothing recalled.
    """
    import time
    import traceback

    while stop is None or not stop.is_set():
        try:
            for result in poll_once(db_path=db_path):
                print(f"[sftp-drop] {result['status']}: {result['filename']}"
                      + (f" -- {result.get('reason', '')}" if result["status"] == "rejected" else ""))
        except Exception:                      # noqa: BLE001
            traceback.print_exc()
        time.sleep(interval_seconds)
